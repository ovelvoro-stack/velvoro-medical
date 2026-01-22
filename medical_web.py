from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook, Workbook
from datetime import datetime
import medical_engine
import os

app = Flask(__name__)

EXCEL_FILE = "medical_records.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time", "Name", "Phone", "Age", "Gender", "Area",
            "Complaint", "Conditions", "Confidence",
            "Protocols", "Medicines", "Tests", "Status"
        ])
        wb.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def patient_input():
    if request.method == "POST":
        patient = {
            "name": request.form["name"],
            "phone": request.form["phone"],
            "age": request.form["age"],
            "gender": request.form["gender"],
            "area": request.form["area"],
            "complaint": request.form["complaint"],
            "symptoms": request.form["symptoms"],
            "sbp": request.form["sbp"],
            "dbp": request.form["dbp"],
            "hr": request.form["hr"]
        }

        result = medical_engine.velvoro_engine(patient)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            patient["name"],
            patient["phone"],
            patient["age"],
            patient["gender"],
            patient["area"],
            patient["complaint"],
            result["conditions"],
            result["confidences"],
            result["protocols"],
            result["medicines"],
            result["tests"],
            "Pending Doctor Review"
        ])
        wb.save(EXCEL_FILE)

        return redirect("/doctor")

    return render_template("patient_input.html")

@app.route("/doctor")
def doctor_panel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template("doctor_panel.html", rows=rows)

if __name__ == "__main__":
    init_excel()
    app.run(debug=True)