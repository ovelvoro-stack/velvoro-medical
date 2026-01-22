from flask import Flask, request, render_template_string
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = "patients_data.xlsx"

# ------------------ Excel Setup ------------------
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Date", "Doctor", "Patient", "Age", "Phone",
        "BP", "Sugar", "Sleep", "Stress", "Symptoms",
        "AI Conditions", "AI Medicine Draft", "Doctor Decision"
    ])
    wb.save(EXCEL_FILE)

# ------------------ AI Decision Engine ------------------
def ai_decision(bp, sugar, sleep, stress):
    conditions = []
    medicines = []

    # BP logic
    try:
        sys, dia = map(int, bp.split("/"))
        if sys >= 140 or dia >= 90:
            conditions.append("High BP")
            medicines.append("Reduce salt, BP monitoring, consult physician")
        elif sys >= 130:
            conditions.append("Borderline BP")
            medicines.append("Daily walking, BP monitoring")
    except:
        conditions.append("Invalid BP")
        medicines.append("Enter BP correctly (example 120/80)")

    # Sugar logic
    try:
        sugar = int(sugar)
        if sugar >= 126:
            conditions.append("Diabetes")
            medicines.append("Metformin 500mg once daily")
        elif sugar >= 100:
            conditions.append("Pre-Diabetes")
            medicines.append("Low sugar diet, exercise")
    except:
        pass

    # Sleep / Stress
    if sleep == "Yes":
        conditions.append("Sleep Disorder")
        medicines.append("Sleep hygiene, reduce screen time")

    if stress == "Yes":
        conditions.append("Stress Related Issue")
        medicines.append("Yoga, meditation, stress management")

    if not conditions:
        conditions.append("No major risk detected")
        medicines.append("Healthy lifestyle")

    return conditions, medicines

# ------------------ Main Route ------------------
@app.route("/", methods=["GET", "POST"])
def index():
    ai_conditions = []
    ai_medicines = []
    doctor_decision = ""

    if request.method == "POST":
        doctor = request.form.get("doctor")
        name = request.form.get("name")
        age = request.form.get("age")
        phone = request.form.get("phone")
        bp = request.form.get("bp")
        sugar = request.form.get("sugar")
        sleep = request.form.get("sleep")
        stress = request.form.get("stress")
        symptoms = request.form.get("symptoms")
        action = request.form.get("action")

        ai_conditions, ai_medicines = ai_decision(bp, sugar, sleep, stress)

        if action:
            doctor_decision = action

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                doctor, name, age, phone,
                bp, sugar, sleep, stress, symptoms,
                ", ".join(ai_conditions),
                ", ".join(ai_medicines),
                doctor_decision
            ])
            wb.save(EXCEL_FILE)

    return render_template_string("""
<!DOCTYPE html>
<html>
<head>
    <title>Velvoro Medical AI</title>
</head>
<body>
<h2>New Patient – Clinical Entry</h2>

<form method="post">
Doctor Name:<br>
<input name="doctor" required><br><br>

Patient Name:<br>
<input name="name" required><br><br>

Age:<br>
<input name="age"><br><br>

Phone:<br>
<input name="phone"><br><br>

BP (example 120/80):<br>
<input name="bp"><br><br>

Sugar:<br>
<input name="sugar"><br><br>

Sleep problem?<br>
<select name="sleep">
<option>No</option>
<option>Yes</option>
</select><br><br>

Stress?<br>
<select name="stress">
<option>No</option>
<option>Yes</option>
</select><br><br>

Symptoms:<br>
<textarea name="symptoms"></textarea><br><br>

<button type="submit">Run AI Draft</button>

{% if ai_conditions %}
<hr>
<h3>AI Draft (For Doctor Review)</h3>

<b>Possible Conditions:</b>
<ul>
{% for c in ai_conditions %}
<li>{{c}}</li>
{% endfor %}
</ul>

<b>AI Suggested Medicines (Draft):</b>
<ul>
{% for m in ai_medicines %}
<li>{{m}}</li>
{% endfor %}
</ul>

<button name="action" value="Approved">Approve</button>
<button name="action" value="Rejected">Reject</button>

<p style="color:red">
AI is advisory only. Final decision by registered doctor.
</p>

{% if doctor_decision %}
<p style="color:green">
Doctor Decision: {{doctor_decision}}
</p>
{% endif %}
{% endif %}
</form>
</body>
</html>
""",
    ai_conditions=ai_conditions,
    ai_medicines=ai_medicines,
    doctor_decision=doctor_decision
    )

# ------------------ Run ------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)